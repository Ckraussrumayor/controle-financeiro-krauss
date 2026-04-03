import streamlit as st
import traceback as _tb

# ─── Guard global: qualquer exceção não-tratada aparece na tela ──────────────
def _fatal(msg: str, exc: Exception):
    """Exibe erro fatal e para o app (garante visibilidade no Streamlit Cloud)."""
    try:
        st.set_page_config(page_title="Erro – Controle Financeiro", page_icon="❌")
    except Exception:
        pass
    st.error(f"❌ {msg}")
    st.code(_tb.format_exc())
    st.stop()

try:
    import pandas as pd
    import calendar
    import re
    import time
    from datetime import datetime
except Exception as _e:
    _fatal("Erro ao importar dependências padrão", _e)

# ─── Constantes de segurança ─────────────────────────────────────────────────
_SESSION_TIMEOUT_MIN = 30       # minutos de inatividade para expirar sessão
_MAX_LOGIN_ATTEMPTS  = 5        # tentativas antes de bloquear
_LOGIN_BLOCK_SEC     = 60       # segundos de bloqueio após excesso de tentativas

try:
    import database as db
    import email_utils
    import auth
except Exception as _e:
    _fatal("Erro ao importar módulos do app", _e)

# ─── Configuração ────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Controle Financeiro Krauss",
    page_icon="💰",
    layout="wide",
    initial_sidebar_state="expanded",
)

MESES_PT = {
    1: "Janeiro", 2: "Fevereiro", 3: "Março", 4: "Abril",
    5: "Maio", 6: "Junho", 7: "Julho", 8: "Agosto",
    9: "Setembro", 10: "Outubro", 11: "Novembro", 12: "Dezembro",
}

try:
    db.init_db()
except Exception as _e:
    _fatal("Erro ao inicializar banco de dados", _e)


# ─── Tela de Login ───────────────────────────────────────────────────────────

def _check_session_timeout():
    """Verifica se a sessão expirou por inatividade."""
    if not st.session_state.get("_autenticado"):
        return
    last = st.session_state.get("_last_activity", 0)
    if last and (time.time() - last) > _SESSION_TIMEOUT_MIN * 60:
        st.session_state["_autenticado"] = False
        st.session_state["_session_expired"] = True
    st.session_state["_last_activity"] = time.time()


def _tela_login():
    """Exibe tela de login. Retorna True se autenticado."""
    _check_session_timeout()
    if st.session_state.get("_autenticado"):
        return True

    modo = st.session_state.get("_login_modo", "login")

    st.markdown(
        "<h1 style='text-align:center'>💰 Controle Financeiro Krauss</h1>",
        unsafe_allow_html=True,
    )

    if st.session_state.pop("_session_expired", False):
        st.warning("⏱️ Sessão expirada por inatividade. Faça login novamente.")

    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        if modo == "login":
            # Rate limiting
            _attempts = st.session_state.get("_login_attempts", 0)
            _blocked_until = st.session_state.get("_login_blocked_until", 0)
            _is_blocked = time.time() < _blocked_until
            if _is_blocked:
                _rest = int(_blocked_until - time.time())
                st.error(f"🔒 Muitas tentativas. Aguarde {_rest}s.")

            with st.form("form_login"):
                st.subheader("🔐 Login")
                usuario = st.text_input("Usuário")
                senha = st.text_input("Senha", type="password")
                entrar = st.form_submit_button("Entrar", use_container_width=True, disabled=_is_blocked)
            if entrar and not _is_blocked:
                if auth.verificar_login(usuario, senha):
                    st.session_state["_autenticado"] = True
                    st.session_state["_last_activity"] = time.time()
                    st.session_state["_login_attempts"] = 0
                    st.session_state.pop("_login_blocked_until", None)
                    st.rerun()
                else:
                    _attempts += 1
                    st.session_state["_login_attempts"] = _attempts
                    if _attempts >= _MAX_LOGIN_ATTEMPTS:
                        st.session_state["_login_blocked_until"] = time.time() + _LOGIN_BLOCK_SEC
                        st.error(f"🔒 Limite de tentativas atingido. Bloqueado por {_LOGIN_BLOCK_SEC}s.")
                    else:
                        restantes = _MAX_LOGIN_ATTEMPTS - _attempts
                        st.error(f"Usuário ou senha incorretos. ({restantes} tentativa(s) restante(s))")

            c1, c2 = st.columns(2)
            with c1:
                if st.button("Esqueci minha senha", use_container_width=True):
                    st.session_state["_login_modo"] = "recuperar"
                    st.rerun()
            with c2:
                if st.button("Redefinir senha", use_container_width=True):
                    st.session_state["_login_modo"] = "redefinir"
                    st.rerun()

        elif modo == "recuperar":
            st.subheader("📧 Recuperar Senha")
            st.info("Um e-mail será enviado para o endereço de recuperação cadastrado.")
            if st.button("Enviar e-mail de recuperação", use_container_width=True):
                resultado = auth.enviar_recuperacao()
                if resultado.startswith("✅"):
                    st.success(resultado)
                else:
                    st.error(resultado)
            if st.button("⬅️ Voltar ao login", use_container_width=True):
                st.session_state["_login_modo"] = "login"
                st.rerun()

        elif modo == "redefinir":
            st.subheader("🔑 Redefinir Senha")
            with st.form("form_redefinir"):
                usuario = st.text_input("Usuário")
                senha_atual = st.text_input("Senha atual", type="password")
                nova_senha = st.text_input("Nova senha", type="password")
                confirmar = st.text_input("Confirmar nova senha", type="password")
                redefinir = st.form_submit_button("Redefinir", use_container_width=True)
            if redefinir:
                if not auth.verificar_login(usuario, senha_atual):
                    st.error("Usuário ou senha atual incorretos.")
                elif len(nova_senha) < 6:
                    st.error("A nova senha deve ter pelo menos 6 caracteres.")
                elif nova_senha != confirmar:
                    st.error("As senhas não coincidem.")
                else:
                    auth.alterar_senha(nova_senha)
                    st.success("✅ Senha alterada com sucesso! Faça login novamente.")
                    st.session_state["_login_modo"] = "login"
            if st.button("⬅️ Voltar ao login", use_container_width=True):
                st.session_state["_login_modo"] = "login"
                st.rerun()

    return False


if not _tela_login():
    st.stop()


# ─── Auto-Sync: restaurar do e-mail ao iniciar a sessão ──────────────────────

def _auto_restore():
    """Restaura backup do e-mail automaticamente na 1ª execução da sessão.
    Só restaura se o banco estiver vazio (ex: deploy acordou do sleep)."""
    if st.session_state.get("_sync_restored"):
        return

    # Se já tem dados locais, não sobrescreve
    if not db.banco_vazio():
        st.session_state["_sync_restored"] = True
        st.session_state["_sync_hash"] = db.db_hash()
        return

    cfg = email_utils.get_config()
    if not email_utils.is_configured(cfg):
        st.session_state["_sync_restored"] = True
        st.session_state["_sync_hash"] = db.db_hash()
        st.session_state["_restore_failed"] = "no_email"
        return

    try:
        excel_bytes, fname = email_utils.get_latest_backup(cfg)
    except Exception as _ex:
        st.session_state["_sync_restored"] = True
        st.session_state["_sync_hash"] = db.db_hash()
        st.session_state["_restore_failed"] = f"erro_imap: {_ex}"
        return

    if excel_bytes is None:
        st.session_state["_sync_restored"] = True
        st.session_state["_sync_hash"] = db.db_hash()
        st.session_state["_restore_failed"] = "no_backup"
        return

    try:
        import io, openpyxl
        db.limpar_banco()
        wb_bk = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)
        _importar_backup_workbook(wb_bk)
        st.toast(f"✅ Dados restaurados do backup: {fname}", icon="📧")
        st.session_state.pop("_restore_failed", None)
    except Exception as e:
        st.toast(f"⚠️ Falha ao restaurar backup: {e}", icon="⚠️")
        st.session_state["_restore_failed"] = f"erro_import: {e}"

    st.session_state["_sync_restored"] = True
    st.session_state["_sync_hash"] = db.db_hash()


def fmt(valor):
    """Formata valor monetário brasileiro."""
    return f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


# ── Helpers de exportação / importação de backup ─────────────────────────────

def _gerar_excel_bytes() -> bytes:
    """Gera o arquivo Excel de backup com todos os dados do banco."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb_out = Workbook()
    hf = Font(bold=True, color="FFFFFF")
    hfill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")

    def _hdr(ws, headers, row=1):
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = hf
            c.fill = hfill
            c.alignment = Alignment(horizontal="center")

    ws1 = wb_out.active
    ws1.title = "Fontes de Renda"
    _hdr(ws1, ["ID", "Nome", "Valor"])
    for f in db.listar_fontes_renda():
        ws1.append([f["id"], f["nome"], f["valor"]])

    ws2 = wb_out.create_sheet("Despesas Planejamento")
    _hdr(ws2, ["ID", "Nome", "Valor", "Detalhes"])
    for d in db.listar_despesas_planejamento():
        ws2.append([d["id"], d["nome"], d["valor"], d["detalhes"]])

    ws3 = wb_out.create_sheet("Lançamentos Mensais")
    _hdr(ws3, ["Ano", "Mês", "Categoria", "Descrição", "Valor"])
    cat_dict_exp = dict(db.CATEGORIAS_MES)
    for m in db.listar_meses():
        for l in db.listar_lancamentos(m["id"]):
            ws3.append([m["ano"], MESES_PT[m["mes"]],
                        cat_dict_exp.get(l["categoria"], l["categoria"]),
                        l["descricao"], l["valor"]])

    ws4 = wb_out.create_sheet("Viagens")
    _hdr(ws4, ["Viagem", "Data", "Categoria", "Descrição", "Valor", "Pago por Nina"])
    cat_dict_v_exp = dict(db.CATEGORIAS_VIAGEM)
    for v in db.listar_viagens():
        for l in db.listar_lancamentos_viagem(v["id"]):
            ws4.append([v["nome"], v["data_viagem"],
                        cat_dict_v_exp.get(l["categoria"], l["categoria"]),
                        l["descricao"], l["valor"],
                        "Sim" if l["pago_por_nina"] else "Não"])

    ws5 = wb_out.create_sheet("Parcelamentos")
    _hdr(ws5, ["Descrição", "Valor Parcela", "Nº Parcelas", "Parcela Atual", "Ativo"])
    for p in db.listar_parcelamentos(ativos=False):
        ws5.append([p["descricao"], p["valor_parcela"], p["num_parcelas"],
                    p["parcela_atual"], "Sim" if p["ativo"] else "Não"])

    ws6 = wb_out.create_sheet("Dívidas Nino")
    _hdr(ws6, ["ID", "Descrição", "Valor Total", "Quitada"])
    for d in db.listar_dividas(apenas_abertas=False):
        ws6.append([d["id"], d["descricao"], d["valor_total"],
                    "Sim" if d["quitada"] else "Não"])

    ws7 = wb_out.create_sheet("Pagamentos Dívidas")
    _hdr(ws7, ["Dívida", "Ano", "Mês", "Valor"])
    all_pgtos = db.listar_pagamentos_divida()
    for pg in all_pgtos:
        m_pg = db.obter_mes(pg["mes_id"])
        if m_pg:
            ws7.append([pg["divida_desc"], m_pg["ano"], MESES_PT[m_pg["mes"]], pg["valor"]])

    for ws in wb_out.worksheets:
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

    buf = io.BytesIO()
    wb_out.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _importar_backup_workbook(wb) -> list:
    """
    Importa dados de um backup Excel gerado pelo próprio app.
    Retorna lista de mensagens de log.
    """
    log = []
    mes_num_map = {v: k for k, v in MESES_PT.items()}
    cat_key_map = {v: k for k, v in dict(db.CATEGORIAS_MES).items()}
    cat_key_map_v = {v: k for k, v in dict(db.CATEGORIAS_VIAGEM).items()}

    if "Fontes de Renda" in wb.sheetnames:
        for row in wb["Fontes de Renda"].iter_rows(min_row=2, values_only=True):
            if row[1] and row[2] is not None:
                db.adicionar_fonte_renda(str(row[1]), float(row[2]))
                log.append(f"✅ Renda: {row[1]}")

    if "Despesas Planejamento" in wb.sheetnames:
        for row in wb["Despesas Planejamento"].iter_rows(min_row=2, values_only=True):
            if row[1] and row[2] is not None:
                db.adicionar_despesa_planejamento(str(row[1]), float(row[2]), row[3])
                log.append(f"✅ Despesa: {row[1]}")

    if "Lançamentos Mensais" in wb.sheetnames:
        meses_cache = {}
        for row in wb["Lançamentos Mensais"].iter_rows(min_row=2, values_only=True):
            ano, mes_nome, cat_nome, desc, valor = (row + (None,) * 5)[:5]
            if not all([ano, mes_nome, cat_nome, desc, valor is not None]):
                continue
            mes_num = mes_num_map.get(mes_nome)
            if not mes_num:
                continue
            key = (int(ano), mes_num)
            if key not in meses_cache:
                m = db.criar_mes(int(ano), mes_num)
                meses_cache[key] = m["id"]
            db.adicionar_lancamento(meses_cache[key],
                                    cat_key_map.get(cat_nome, cat_nome),
                                    str(desc), float(valor))
        log.append("✅ Lançamentos mensais")

    if "Viagens" in wb.sheetnames:
        viagens_cache = {}
        for row in wb["Viagens"].iter_rows(min_row=2, values_only=True):
            v_nome, v_data, cat_nome, desc, valor, nina_str = (list(row) + [None] * 6)[:6]
            if not all([v_nome, cat_nome, desc, valor is not None]):
                continue
            if v_nome not in viagens_cache:
                db.criar_viagem(v_nome, str(v_data) if v_data else None, None)
                todas = db.listar_viagens()
                viagens_cache[v_nome] = next(
                    (v["id"] for v in todas if v["nome"] == v_nome), None
                )
            vid = viagens_cache[v_nome]
            if vid:
                db.adicionar_lancamento_viagem(
                    vid, cat_key_map_v.get(cat_nome, cat_nome),
                    str(desc), float(valor), nina_str == "Sim"
                )
        log.append("✅ Viagens")

    if "Parcelamentos" in wb.sheetnames:
        for row in wb["Parcelamentos"].iter_rows(min_row=2, values_only=True):
            desc, val_p, num_p, p_atual, ativo_str = (list(row) + [None] * 5)[:5]
            if not desc:
                continue
            db.adicionar_parcelamento(str(desc), float(val_p), int(num_p), int(p_atual))
            if ativo_str == "Não":
                for p in db.listar_parcelamentos(ativos=True):
                    if p["descricao"] == desc:
                        db.finalizar_parcelamento(p["id"])
                        break
            log.append(f"✅ Parcelamento: {desc}")

    if "Dívidas Nino" in wb.sheetnames:
        dividas_map = {}  # id original -> novo id
        for row in wb["Dívidas Nino"].iter_rows(min_row=2, values_only=True):
            id_orig, desc, val_total, quitada_str = (list(row) + [None] * 4)[:4]
            if not desc:
                continue
            db.adicionar_divida(str(desc), float(val_total))
            todas = db.listar_dividas(apenas_abertas=False)
            novo = next((d for d in reversed(todas) if d["descricao"] == desc), None)
            if novo:
                dividas_map[id_orig] = novo["id"]
                if quitada_str == "Sim":
                    db.quitar_divida(novo["id"])
        log.append("✅ Dívidas Nino")

    if "Pagamentos Dívidas" in wb.sheetnames:
        mes_num_map2 = {v: k for k, v in MESES_PT.items()}
        meses_cache2 = {}
        for row in wb["Pagamentos Dívidas"].iter_rows(min_row=2, values_only=True):
            div_desc, ano, mes_nome, valor = (list(row) + [None] * 4)[:4]
            if not all([div_desc, ano, mes_nome, valor is not None]):
                continue
            mes_num = mes_num_map2.get(mes_nome)
            if not mes_num:
                continue
            key = (int(ano), mes_num)
            if key not in meses_cache2:
                m2 = db.criar_mes(int(ano), mes_num)
                meses_cache2[key] = m2["id"]
            # encontrar dívida pelo nome
            todas_div = db.listar_dividas(apenas_abertas=False)
            div_match = next((d for d in todas_div if d["descricao"] == div_desc), None)
            if div_match:
                db.adicionar_pagamento_divida(div_match["id"], meses_cache2[key], float(valor))
        log.append("✅ Pagamentos de dívidas")

    return log


def nome_mes(ano, mes):
    return f"{MESES_PT[mes]} {ano}"


# ─── Executar auto-restore na inicialização da sessão ───────────────────────
_auto_restore()

# ─── Aviso persistente se auto-restore falhou e banco está vazio ────────────
_restore_fail = st.session_state.get("_restore_failed")
if _restore_fail and db.banco_vazio():
    if _restore_fail == "no_email":
        st.warning("⚠️ **Banco vazio** — E-mail de backup não configurado. Configure em 📥 Importar / Exportar → ⚙️ Config E-mail, ou importe um backup manualmente.")
    elif _restore_fail == "no_backup":
        st.warning("⚠️ **Banco vazio** — Nenhum backup encontrado no e-mail. Importe dados manualmente em 📥 Importar / Exportar.")
    elif _restore_fail.startswith("erro_imap"):
        st.error(f"⚠️ **Banco vazio** — Falha ao conectar ao e-mail para restaurar: `{_restore_fail.split(': ', 1)[-1]}`. Tente restaurar manualmente.")
    elif _restore_fail.startswith("erro_import"):
        st.error(f"⚠️ **Banco vazio** — Erro ao importar backup: `{_restore_fail.split(': ', 1)[-1]}`. Tente restaurar manualmente.")
    else:
        st.warning(f"⚠️ **Banco vazio** — Restauração automática falhou: {_restore_fail}")


# ─── CSS personalizado ──────────────────────────────────────────────────────
st.markdown("""
<style>
    .metric-card {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 1.2rem; border-radius: 12px; color: white;
        text-align: center; margin-bottom: 0.5rem;
    }
    .metric-card.green { background: linear-gradient(135deg, #11998e 0%, #38ef7d 100%); }
    .metric-card.red { background: linear-gradient(135deg, #eb3349 0%, #f45c43 100%); }
    .metric-card.blue { background: linear-gradient(135deg, #2193b0 0%, #6dd5ed 100%); }
    .metric-card.orange { background: linear-gradient(135deg, #f7971e 0%, #ffd200 100%); }
    .metric-card h3 { margin: 0; font-size: 0.85rem; opacity: 0.9; }
    .metric-card h2 { margin: 0.3rem 0 0 0; font-size: 1.5rem; }
    .cat-header { border-left: 4px solid #667eea; padding-left: 10px; margin: 1rem 0 0.5rem 0; }
    div[data-testid="stSidebar"] { background-color: #1a1a2e; }
</style>
""", unsafe_allow_html=True)


def metric_card(label, value, css_class=""):
    st.markdown(
        f'<div class="metric-card {css_class}"><h3>{label}</h3><h2>{fmt(value)}</h2></div>',
        unsafe_allow_html=True
    )


def _resumo_mes_html(mes_id) -> str:
    """Gera HTML estilo planilha para o resumo mensal (para print/WhatsApp)."""
    m = db.obter_mes(mes_id)
    mes_titulo = nome_mes(m["ano"], m["mes"]).upper()
    totais = db.totais_mes(mes_id)
    pgtos_dividas_total = db.total_pagamentos_mes(mes_id)
    total_mes = sum(totais.values()) + pgtos_dividas_total

    CAT_HDR = 'style="background:#1a1a2e;color:#fff;font-weight:bold;padding:7px 10px;"'
    SUB_TD  = 'style="background:#eef4fb;font-style:italic;padding:3px 10px;border-top:1px dashed #bbb;"'
    DESC_TD = 'style="padding:3px 10px;color:#222;"'
    VAL_TD  = 'style="padding:3px 10px;text-align:right;white-space:nowrap;color:#222;"'
    SUB_VAL = 'style="background:#eef4fb;font-style:italic;padding:3px 10px;text-align:right;border-top:1px dashed #bbb;"'

    cat_icons = {
        "conta_fixa": "🏠", "parcelada": "📅", "desp_nina": "🛒",
        "devedor_nina": "💳", "extras_nino": "🎁"
    }

    linhas_html = ""
    for cat_key, cat_nome in db.CATEGORIAS_MES:
        lancamentos = db.listar_lancamentos(mes_id, cat_key)
        subtotal = totais.get(cat_key, 0)
        icon = cat_icons.get(cat_key, "📌")

        if cat_key == "extras_nino":
            pgtos_mes = db.listar_pagamentos_divida(mes_id=mes_id)
            pgtos_validos = [p for p in pgtos_mes if p["valor"] > 0.01]
            if not lancamentos and not pgtos_validos:
                continue
            linhas_html += f'<tr><td colspan="2" {CAT_HDR}>{icon} {cat_nome}</td></tr>'
            for l in lancamentos:
                linhas_html += f'<tr><td {DESC_TD}>&nbsp;&nbsp;{l["descricao"]}</td><td {VAL_TD}>{fmt(l["valor"])}</td></tr>'
            for p in pgtos_validos:
                saldo = db.saldo_divida(p["divida_id"])
                saldo_span = f'<span style="font-size:0.78rem;color:#666;">(saldo: {fmt(saldo)})</span>'
                linhas_html += f'<tr><td {DESC_TD}>&nbsp;&nbsp;💳 {p["divida_desc"]} {saldo_span}</td><td {VAL_TD}>{fmt(p["valor"])}</td></tr>'
            subtotal_real = subtotal + db.total_pagamentos_mes(mes_id)
            linhas_html += f'<tr><td {SUB_TD}>&nbsp;&nbsp;Subtotal</td><td {SUB_VAL}>{fmt(subtotal_real)}</td></tr>'
        else:
            if not lancamentos and subtotal == 0:
                continue
            linhas_html += f'<tr><td colspan="2" {CAT_HDR}>{icon} {cat_nome}</td></tr>'
            for l in lancamentos:
                linhas_html += f'<tr><td {DESC_TD}>&nbsp;&nbsp;{l["descricao"]}</td><td {VAL_TD}>{fmt(l["valor"])}</td></tr>'
            linhas_html += f'<tr><td {SUB_TD}>&nbsp;&nbsp;Subtotal</td><td {SUB_VAL}>{fmt(subtotal)}</td></tr>'

    devolver = totais.get("desp_nina", 0) + totais.get("conta_fixa", 0) + totais.get("parcelada", 0) + totais.get("extras_nino", 0) + pgtos_dividas_total
    creditos = totais.get("devedor_nina", 0)
    liquido_nina = devolver + creditos

    import math
    arredondado = math.ceil(total_mes / 50) * 50

    html = f"""<div style="font-family:'Courier New',monospace;background:#fff;color:#111;
border:2px solid #333;border-radius:8px;padding:20px;max-width:520px;margin:0 auto;">
<div style="text-align:center;border-bottom:2px solid #333;padding-bottom:10px;margin-bottom:12px;">
  <div style="font-size:1.3rem;font-weight:bold;">💰 ACERTO MENSAL</div>
  <div style="font-size:1.1rem;">{mes_titulo}</div>
</div>
<table style="width:100%;border-collapse:collapse;font-size:0.93rem;">
{linhas_html}
<tr><td colspan="2" style="padding:6px;"></td></tr>
<tr style="background:#2193b0;color:#fff;font-size:1.05rem;">
  <td style="padding:8px;font-weight:bold;">TOTAL DO MÊS</td>
  <td style="padding:8px;text-align:right;font-weight:bold;">{fmt(total_mes)}</td>
</tr>
<tr style="background:#f7971e;color:#fff;">
  <td style="padding:6px 8px;">Arredondado (↑50)</td>
  <td style="padding:6px 8px;text-align:right;">{fmt(arredondado)}</td>
</tr>
<tr style="background:#11998e;color:#fff;">
  <td style="padding:6px 8px;">Líquido Nina</td>
  <td style="padding:6px 8px;text-align:right;">{fmt(liquido_nina)}</td>
</tr>
</table>
{f'<div style="margin-top:10px;padding:8px 10px;background:#f5f5f5;border-radius:6px;font-size:0.85rem;color:#555;"><strong>\U0001f4dd Obs:</strong> {m["observacoes"]}</div>' if m["observacoes"] else ''}
<div style="text-align:center;margin-top:12px;font-size:0.75rem;color:#999;">
  Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}
</div>
</div>"""
    return html


def _resumo_viagem_html(viagem_id) -> str:
    """Gera HTML estilo planilha para resumo de viagem (para print/WhatsApp)."""
    v = db.obter_viagem(viagem_id)
    titulo = v["nome"].upper()
    data_str = v["data_viagem"] or ""
    if data_str:
        try:
            data_str = datetime.strptime(data_str, "%Y-%m-%d").strftime("%d/%m/%Y")
        except ValueError:
            pass
    totais_v = db.totais_viagem(viagem_id)
    total_v = db.total_viagem(viagem_id)
    total_nina = db.total_nina_viagem(viagem_id)

    CAT_HDR = 'style="background:#1a1a2e;color:#fff;font-weight:bold;padding:7px 10px;"'
    SUB_TD  = 'style="background:#eef4fb;font-style:italic;padding:3px 10px;border-top:1px dashed #bbb;"'
    DESC_TD = 'style="padding:3px 10px;color:#222;"'
    VAL_TD  = 'style="padding:3px 10px;text-align:right;white-space:nowrap;color:#222;"'
    SUB_VAL = 'style="background:#eef4fb;font-style:italic;padding:3px 10px;text-align:right;border-top:1px dashed #bbb;"'

    cat_icons_v = {
        "pedagio": "🚗", "combustivel": "⛽",
        "despesa_viagem": "🍽️", "desp_nina_viagem": "🛒",
    }

    linhas_html = ""
    for cat_key, cat_nome in db.CATEGORIAS_VIAGEM:
        lancamentos = db.listar_lancamentos_viagem(viagem_id, cat_key)
        subtotal = totais_v.get(cat_key, 0)
        if not lancamentos and subtotal == 0:
            continue
        icon = cat_icons_v.get(cat_key, "📌")
        linhas_html += f'<tr><td colspan="3" {CAT_HDR}>{icon} {cat_nome}</td></tr>'
        for l in lancamentos:
            nina_tag = '<span style="font-size:0.75rem;background:#f7971e;color:#fff;padding:1px 5px;border-radius:3px;">Nina</span>' if l["pago_por_nina"] else ""
            linhas_html += f'<tr><td {DESC_TD}>&nbsp;&nbsp;{l["descricao"]}</td><td {VAL_TD}>{fmt(l["valor"])}</td><td {VAL_TD}>{nina_tag}</td></tr>'
        linhas_html += f'<tr><td colspan="2" {SUB_TD}>&nbsp;&nbsp;Subtotal</td><td {SUB_VAL}>{fmt(subtotal)}</td></tr>'

    html = f"""<div style="font-family:'Courier New',monospace;background:#fff;color:#111;
border:2px solid #333;border-radius:8px;padding:20px;max-width:560px;margin:0 auto;">
<div style="text-align:center;border-bottom:2px solid #333;padding-bottom:10px;margin-bottom:12px;">
  <div style="font-size:1.3rem;font-weight:bold;">✈️ RESUMO DA VIAGEM</div>
  <div style="font-size:1.1rem;">{titulo}</div>
  <div style="font-size:0.85rem;color:#555;">{data_str}</div>
</div>
<table style="width:100%;border-collapse:collapse;font-size:0.93rem;">
{linhas_html}
<tr><td colspan="3" style="padding:6px;"></td></tr>
<tr style="background:#2193b0;color:#fff;font-size:1.05rem;">
  <td colspan="2" style="padding:8px;font-weight:bold;">TOTAL VIAGEM</td>
  <td style="padding:8px;text-align:right;font-weight:bold;">{fmt(total_v)}</td>
</tr>
<tr style="background:#f7971e;color:#fff;">
  <td colspan="2" style="padding:6px 8px;">Pago por Nina</td>
  <td style="padding:6px 8px;text-align:right;">{fmt(total_nina)}</td>
</tr>
<tr style="background:#eb3349;color:#fff;">
  <td colspan="2" style="padding:6px 8px;">Meu Custo</td>
  <td style="padding:6px 8px;text-align:right;">{fmt(total_v - total_nina)}</td>
</tr>
</table>
{f'<div style="margin-top:10px;padding:8px 10px;background:#f5f5f5;border-radius:6px;font-size:0.85rem;color:#555;"><strong>\U0001f4dd Obs:</strong> {v["observacoes"]}</div>' if ("observacoes" in v.keys() and v["observacoes"]) else ''}
<div style="text-align:center;margin-top:12px;font-size:0.75rem;color:#999;">
  Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}
</div>
</div>"""
    return html


# ═══════════════════════════════════════════════════════════════════════════
# SIDEBAR – Navegação
# ═══════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.title("💰 Financeiro Krauss")
    pagina = st.radio(
        "Navegação",
        [
            "🏠 Visão Geral",
            "📋 Planejamento Mensal",
            "📝 Contas do Mês",
            "✈️ Viagens / Eventos",
            "📊 Histórico",
            "📦 Parcelamentos",
            "📥 Importar / Exportar",
        ],
        label_visibility="collapsed",
    )
    st.divider()
    # Indicador de backup pendente
    if st.session_state.get("_sync_hash") and db.db_hash() != st.session_state.get("_sync_hash", ""):
        st.warning("⚠️ Dados alterados\nEncerre a sessão para enviar o backup.", icon="☁️")
    if st.button("⏹️ Encerrar Sessão", use_container_width=True):
        # Backup final antes de fechar
        cfg_exit = email_utils.get_config()
        _backup_ok = True
        if email_utils.is_configured(cfg_exit):
            current_h = db.db_hash()
            stored_h = st.session_state.get("_sync_hash")
            if current_h != stored_h:
                try:
                    excel_b = _gerar_excel_bytes()
                    fn = f"financeiro_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                    email_utils.send_backup(excel_b, fn, cfg_exit)
                    email_utils.delete_old_backups(cfg_exit)
                    st.toast("☁️ Backup final enviado", icon="✅")
                except Exception as _bk_err:
                    _backup_ok = False
                    st.session_state["_backup_fail"] = str(_bk_err)
        if _backup_ok or st.session_state.get("_force_logout"):
            st.session_state.clear()
            st.info("Sessão encerrada. Recarregue a página para entrar novamente.")
            st.stop()

    # Mostrar erro de backup e opção de forçar logout
    if st.session_state.get("_backup_fail"):
        st.error(f"❌ Falha ao enviar backup: {st.session_state['_backup_fail']}")
        c_retry, c_force = st.columns(2)
        with c_retry:
            if st.button("🔄 Tentar novamente", key="retry_bk", use_container_width=True):
                st.session_state.pop("_backup_fail", None)
                st.rerun()
        with c_force:
            if st.button("⚠️ Sair sem backup", key="force_logout", use_container_width=True):
                st.session_state["_force_logout"] = True
                st.rerun()
    st.caption(f"v1.0 • Controle Financeiro • {datetime.now().strftime('%d/%m/%Y')}")


# ═══════════════════════════════════════════════════════════════════════════
# 🏠 VISÃO GERAL
# ═══════════════════════════════════════════════════════════════════════════
if pagina == "🏠 Visão Geral":
    st.header("🏠 Visão Geral")

    total_r = db.total_renda()
    total_d = db.total_despesas_planejamento()
    saldo = total_r - total_d

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        metric_card("Receita Total", total_r, "green")
    with c2:
        metric_card("Despesas Planejadas", total_d, "red")
    with c3:
        metric_card("Saldo Mensal", saldo, "blue" if saldo >= 0 else "red")
    with c4:
        metric_card("Sobra Anual (12 meses)", saldo * 12, "orange")

    st.divider()

    # Últimos meses
    meses = db.listar_meses()
    if meses:
        st.subheader("📅 Últimos Acertos Mensais (Nina)")
        dados_hist = []
        for m in meses[:6]:
            tot = db.total_geral_mes(m["id"]) + db.total_pagamentos_mes(m["id"])
            dados_hist.append({
                "Mês": nome_mes(m["ano"], m["mes"]),
                "Total Acerto": tot,
            })
        df = pd.DataFrame(dados_hist)
        st.dataframe(df.style.format({"Total Acerto": "R$ {:.2f}"}), use_container_width=True, hide_index=True)

        # Gráfico
        if len(dados_hist) > 1:
            import plotly.express as px
            _df_bar = pd.DataFrame(dados_hist)
            _fig_bar = px.bar(_df_bar, x="Mês", y="Total Acerto",
                              color_discrete_sequence=["#667eea"],
                              labels={"Total Acerto": "R$"})
            _fig_bar.update_layout(
                xaxis_tickangle=-45,
                xaxis_title=None,
                yaxis_title="R$",
                showlegend=False,
                height=340,
                margin=dict(b=80),
            )
            _fig_bar.update_traces(hovertemplate="%{x}<br>R$ %{y:,.2f}<extra></extra>")
            st.plotly_chart(_fig_bar, use_container_width=True)

        # ── Dashboard consolidado ──
        st.divider()
        st.subheader("📊 Dashboard Consolidado")

        todos_totais = []
        for m in meses:
            todos_totais.append(db.total_geral_mes(m["id"]) + db.total_pagamentos_mes(m["id"]))

        media_mensal = sum(todos_totais) / len(todos_totais) if todos_totais else 0
        mes_mais_caro_idx = todos_totais.index(max(todos_totais)) if todos_totais else 0
        mes_mais_barato_idx = todos_totais.index(min(todos_totais)) if todos_totais else 0

        c1, c2, c3, c4 = st.columns(4)
        with c1:
            metric_card("Média Mensal", media_mensal, "blue")
        with c2:
            metric_card("Mês + Caro", max(todos_totais) if todos_totais else 0, "red")
        with c3:
            metric_card("Mês + Barato", min(todos_totais) if todos_totais else 0, "green")
        with c4:
            metric_card("Total Acumulado", sum(todos_totais), "orange")

        if todos_totais:
            st.caption(
                f"🔴 Mais caro: **{nome_mes(meses[mes_mais_caro_idx]['ano'], meses[mes_mais_caro_idx]['mes'])}** · "
                f"🟢 Mais barato: **{nome_mes(meses[mes_mais_barato_idx]['ano'], meses[mes_mais_barato_idx]['mes'])}**"
            )

        # Comparativo planejado vs real
        if total_d > 0 and todos_totais:
            st.divider()
            st.subheader("📐 Planejado vs Real")
            comp_data = []
            for i, m in enumerate(meses):
                comp_data.append({
                    "Mês": nome_mes(m["ano"], m["mes"]),
                    "Planejado": total_d,
                    "Real": todos_totais[i],
                })
            import plotly.express as px
            _df_comp_long = pd.DataFrame(comp_data).melt(id_vars="Mês", var_name="Tipo", value_name="Valor")
            _fig_comp = px.bar(
                _df_comp_long, x="Mês", y="Valor", color="Tipo", barmode="group",
                color_discrete_map={"Planejado": "#667eea", "Real": "#f7971e"},
                labels={"Valor": "R$"},
            )
            _fig_comp.update_layout(
                xaxis_tickangle=-45,
                xaxis_title=None,
                yaxis_title="R$",
                height=340,
                margin=dict(b=80),
            )
            _fig_comp.update_traces(hovertemplate="%{x}<br>R$ %{y:,.2f}<extra></extra>")
            st.plotly_chart(_fig_comp, use_container_width=True)
    else:
        st.info("Nenhum mês registrado ainda. Vá em **Contas do Mês** para criar o primeiro.")

    # Viagens recentes
    viagens = db.listar_viagens()
    if viagens:
        st.subheader("✈️ Últimas Viagens")
        for v in viagens[:3]:
            tot = db.total_viagem(v["id"])
            nina_v = db.total_nina_viagem(v["id"])
            status = " ✅" if ("pago" in v.keys() and v["pago"]) else ""
            st.markdown(f"**{v['nome']}**{status} — Total: {fmt(tot)} · Meu custo: {fmt(tot - nina_v)}")


# ═══════════════════════════════════════════════════════════════════════════
# 📋 PLANEJAMENTO MENSAL
# ═══════════════════════════════════════════════════════════════════════════
elif pagina == "📋 Planejamento Mensal":
    st.header("📋 Planejamento Mensal")
    st.caption("Projeção conservadora de quanto precisa ganhar por mês para cobrir todos os gastos.")

    tab_renda, tab_desp = st.tabs(["💵 Fontes de Renda", "💸 Despesas Planejadas"])

    # --- Fontes de Renda ---
    with tab_renda:
        fontes = db.listar_fontes_renda()
        total_r = db.total_renda()

        col_a, col_b = st.columns([3, 1])
        with col_b:
            metric_card("Total Renda", total_r, "green")

        with col_a:
            if fontes:
                for f in fontes:
                    c1, c2, c3 = st.columns([3, 2, 1])
                    with c1:
                        novo_nome = st.text_input("Nome", f["nome"], key=f"rn_{f['id']}", label_visibility="collapsed")
                    with c2:
                        novo_val = st.number_input("Valor", value=float(f["valor"]), step=50.0, key=f"rv_{f['id']}", label_visibility="collapsed")
                    with c3:
                        with st.popover("🗑️"):
                            st.caption(f"Excluir **{f['nome']}**?")
                            if st.button("✅ Confirmar", key=f"rd_ok_{f['id']}", type="primary"):
                                db.remover_fonte_renda(f["id"])
                                st.rerun()
                    if novo_nome != f["nome"] or novo_val != f["valor"]:
                        db.atualizar_fonte_renda(f["id"], novo_nome, novo_val)
                        st.rerun()

        st.divider()
        with st.expander("➕ Adicionar Fonte de Renda"):
            with st.form("add_renda", clear_on_submit=True):
                c1, c2 = st.columns(2)
                with c1:
                    nome = st.text_input("Nome da fonte")
                with c2:
                    valor = st.number_input("Valor (R$)", value=0.0, min_value=0.0, step=50.0, format="%.2f")
                if st.form_submit_button("Adicionar", type="primary"):
                    if not nome:
                        st.warning("Informe o nome da fonte.")
                    elif valor <= 0:
                        st.warning("Informe um valor maior que zero.")
                    else:
                        db.adicionar_fonte_renda(nome, valor)
                        st.rerun()

    # --- Despesas Planejadas ---
    with tab_desp:
        despesas = db.listar_despesas_planejamento()
        total_d = db.total_despesas_planejamento()

        col_a, col_b = st.columns([3, 1])
        with col_b:
            metric_card("Total Despesas", total_d, "red")
            saldo = db.total_renda() - total_d
            metric_card("Saldo", saldo, "green" if saldo >= 0 else "red")

        with col_a:
            if despesas:
                for d in despesas:
                    c1, c2, c3, c4 = st.columns([3, 2, 2, 1])
                    with c1:
                        novo_nome = st.text_input("Nome", d["nome"], key=f"dn_{d['id']}", label_visibility="collapsed")
                    with c2:
                        novo_val = st.number_input("Valor", value=float(d["valor"]), step=50.0, key=f"dv_{d['id']}", label_visibility="collapsed")
                    with c3:
                        novo_det = st.text_input("Detalhes", d["detalhes"] or "", key=f"dd_{d['id']}", label_visibility="collapsed", placeholder="detalhes...")
                    with c4:
                        with st.popover("🗑️"):
                            st.caption(f"Excluir **{d['nome']}**?")
                            if st.button("✅ Confirmar", key=f"dx_ok_{d['id']}", type="primary"):
                                db.remover_despesa_planejamento(d["id"])
                                st.rerun()
                    if novo_nome != d["nome"] or novo_val != d["valor"] or (novo_det or None) != d["detalhes"]:
                        db.atualizar_despesa_planejamento(d["id"], novo_nome, novo_val, novo_det or None)
                        st.rerun()

        st.divider()
        with st.expander("➕ Adicionar Despesa Planejada"):
            with st.form("add_desp", clear_on_submit=True):
                c1, c2, c3 = st.columns(3)
                with c1:
                    nome = st.text_input("Nome da despesa")
                with c2:
                    valor = st.number_input("Valor (R$)", value=0.0, min_value=0.0, step=50.0, format="%.2f")
                with c3:
                    detalhes = st.text_input("Detalhes (opcional)")
                if st.form_submit_button("Adicionar", type="primary"):
                    if not nome:
                        st.warning("Informe o nome da despesa.")
                    elif valor <= 0:
                        st.warning("Informe um valor maior que zero.")
                    else:
                        db.adicionar_despesa_planejamento(nome, valor, detalhes or None)
                        st.rerun()


# ═══════════════════════════════════════════════════════════════════════════
# 📝 CONTAS DO MÊS
# ═══════════════════════════════════════════════════════════════════════════
elif pagina == "📝 Contas do Mês":
    st.header("📝 Contas do Mês — Acerto com Nina")

    meses = db.listar_meses()
    hoje = datetime.now()

    # Criar novo mês
    with st.expander("➕ Criar Novo Mês"):
        _c1, _c2 = st.columns(2)
        with _c1:
            novo_mes = st.selectbox("Mês", range(1, 13), format_func=lambda x: MESES_PT[x],
                                    index=hoje.month - 1)
        with _c2:
            novo_ano = st.number_input("Ano", min_value=2020, max_value=2040, value=hoje.year)

        _copiar = st.checkbox("📋 Copiar lançamentos fixos de outro mês?")
        _mes_origem_id = None
        _cats_sel = []
        if _copiar and meses:
            _mes_origem_opts = {_m["id"]: nome_mes(_m["ano"], _m["mes"]) for _m in meses}
            _mes_origem_id = st.selectbox(
                "Mês de origem",
                options=list(_mes_origem_opts.keys()),
                format_func=lambda x: _mes_origem_opts[x],
                key="novo_mes_origem_copia"
            )
            st.caption("Selecione o que copiar:")
            _ca, _cb, _cc = st.columns(3)
            with _ca:
                if st.checkbox("🏠 Contas Fixas", value=True, key="copy_cat_conta_fixa"):
                    _cats_sel.append("conta_fixa")
            with _cb:
                if st.checkbox("💳 Parceladas", value=False, key="copy_cat_parcelada"):
                    _cats_sel.append("parcelada")
            with _cc:
                if st.checkbox("👤 Devedor (Nina)", value=True, key="copy_cat_devedor_nina"):
                    _cats_sel.append("devedor_nina")

        if st.button("✅ Criar Mês", type="primary"):
            _novo_m = db.criar_mes(int(novo_ano), novo_mes)
            if _copiar and _mes_origem_id and _cats_sel:
                _m_orig_info = db.obter_mes(_mes_origem_id)
                _nome_orig_up = MESES_PT[_m_orig_info["mes"]].upper()
                _nome_dest_up = MESES_PT[novo_mes].upper()
                _parc_ativos_map = {p["descricao"]: p for p in db.listar_parcelamentos(ativos=True)}
                _parc_avancados = set()

                for _cat in _cats_sel:
                    for _l in db.listar_lancamentos(_mes_origem_id, _cat):
                        _desc = _l["descricao"]
                        if _cat == "parcelada":
                            # Remove numeração anterior (ex: "Seguro (3/10)" → "Seguro")
                            _base = re.sub(r'\s*\(\d+/\d+\)$', '', _desc).strip()
                            _pm = _parc_ativos_map.get(_base)
                            if _pm and _pm["id"] not in _parc_avancados:
                                _prox = min(_pm["parcela_atual"] + 1, _pm["num_parcelas"])
                                _desc = f"{_pm['descricao']} ({_prox}/{_pm['num_parcelas']})"
                                db.atualizar_parcelamento(
                                    _pm["id"], _pm["descricao"],
                                    _pm["valor_parcela"], _pm["num_parcelas"], _prox
                                )
                                if _prox >= _pm["num_parcelas"]:
                                    db.finalizar_parcelamento(_pm["id"])
                                _parc_avancados.add(_pm["id"])
                        elif _cat == "devedor_nina":
                            _desc = _desc.replace(_nome_orig_up, _nome_dest_up)
                        db.adicionar_lancamento(_novo_m["id"], _cat, _desc, _l["valor"])
            st.rerun()

    if not meses:
        st.info("Crie o primeiro mês acima para começar.")
        st.stop()

    # Selecionar mês — persiste ao navegar entre páginas
    opcoes_mes = {
        m["id"]: nome_mes(m["ano"], m["mes"]) + (" ✅ Pago" if ("pago" in m.keys() and m["pago"]) else "")
        for m in meses
    }
    ids_disponiveis = list(opcoes_mes.keys())
    _mes_salvo = st.session_state.get("_mes_selecionado")
    idx_default = ids_disponiveis.index(_mes_salvo) if _mes_salvo in ids_disponiveis else 0
    mes_selecionado = st.selectbox(
        "Selecione o mês",
        options=ids_disponiveis,
        format_func=lambda x: opcoes_mes[x],
        index=idx_default,
    )
    st.session_state["_mes_selecionado"] = mes_selecionado

    m = db.obter_mes(mes_selecionado)
    _pago_mes = bool(m["pago"]) if m and "pago" in m.keys() else False

    if _pago_mes:
        st.markdown("""
        <style>
        section[data-testid="stMain"] > div {
            background: rgba(40, 167, 90, 0.06) !important;
        }
        </style>""", unsafe_allow_html=True)
        st.success("✅ **MÊS PAGO** — Acerto concluído. Edições desativadas.")
        if st.button("🔓 Desmarcar como Pago", key="des_pago_mes"):
            db.marcar_mes_pago(mes_selecionado, 0)
            st.rerun()
    else:
        if st.button("✅ Marcar como Pago", key="marcar_pago_mes", type="secondary",
                     help="Bloqueia edições e sinaliza o mês como acertado"):
            db.marcar_mes_pago(mes_selecionado, 1)
            st.rerun()

    totais = db.totais_mes(mes_selecionado)
    pgtos_dividas_total = db.total_pagamentos_mes(mes_selecionado)

    # Resumo
    total_mes = sum(totais.values()) + pgtos_dividas_total
    import math
    arredondado = math.ceil(total_mes / 50) * 50

    c1, c2, c3 = st.columns(3)
    with c1:
        metric_card("Total do Mês", total_mes, "blue")
    with c2:
        metric_card("Arredondado (↑50)", arredondado, "orange")
    with c3:
        # Subtotais
        devolver = totais.get("desp_nina", 0) + totais.get("conta_fixa", 0) + totais.get("parcelada", 0) + totais.get("extras_nino", 0) + pgtos_dividas_total
        creditos = totais.get("devedor_nina", 0)
        metric_card("Líquido Nina", devolver + creditos, "green" if devolver + creditos >= 0 else "red")

    # Observações do mês
    obs_atual = m["observacoes"] or ""
    nova_obs = st.text_input(
        "📝 Observações do mês (opcional)",
        value=obs_atual,
        placeholder="Ex: Mês de férias, aniversário, despesa extra pontual...",
        key=f"obs_mes_{mes_selecionado}",
        disabled=_pago_mes,
    )
    if not _pago_mes and nova_obs != obs_atual:
        db.atualizar_observacoes_mes(mes_selecionado, nova_obs)
        st.rerun()

    st.divider()

    # Lançamentos por categoria
    cat_dict = dict(db.CATEGORIAS_MES)
    cat_icons = {
        "conta_fixa": "🏠", "parcelada": "📅", "desp_nina": "🛒",
        "devedor_nina": "💳", "extras_nino": "🎁"
    }

    for cat_key, cat_nome in db.CATEGORIAS_MES:
        icon = cat_icons.get(cat_key, "📌")
        lancamentos = db.listar_lancamentos(mes_selecionado, cat_key)
        subtotal = totais.get(cat_key, 0)

        # ── Extras (Nino): seção especial com dívidas ──
        if cat_key == "extras_nino":
            total_pgtos_dividas = db.total_pagamentos_mes(mes_selecionado)
            subtotal_extras = subtotal + total_pgtos_dividas

            st.markdown(f'<div class="cat-header"><strong>{icon} {cat_nome}</strong> — {fmt(subtotal_extras)}</div>',
                        unsafe_allow_html=True)

            # --- Pagamentos de dívidas neste mês ---
            pgtos_mes = db.listar_pagamentos_divida(mes_id=mes_selecionado)
            if pgtos_mes:
                st.caption("💳 Pagamentos de dívidas neste mês:")
                for pg in pgtos_mes:
                    saldo = db.saldo_divida(pg["divida_id"])
                    c1, c2, c3 = st.columns([4, 2, 1])
                    with c1:
                        st.text(f"↳ {pg['divida_desc']}  (saldo: {fmt(saldo)})")
                    with c2:
                        st.text(fmt(pg["valor"]))
                    with c3:
                        if not _pago_mes:
                            with st.popover("🗑️"):
                                st.caption(f"Excluir pagamento de **{pg['divida_desc']}**?")
                                if st.button("✅ Confirmar", key=f"pgd_ok_{pg['id']}", type="primary"):
                                    db.remover_pagamento_divida(pg["id"])
                                    st.rerun()

            if not _pago_mes:
                # --- Selecionar dívida e pagar ---
                dividas_abertas = db.listar_dividas(apenas_abertas=True)
                if dividas_abertas:
                    st.markdown("**💰 Registrar pagamento de dívida:**")
                    opcoes_div = {d["id"]: f"{d['descricao']} — saldo {fmt(db.saldo_divida(d['id']))}"
                                  for d in dividas_abertas}
                    with st.form(f"pagar_divida_{mes_selecionado}", clear_on_submit=True):
                        fc1, fc2, fc3 = st.columns([4, 2, 1])
                        with fc1:
                            div_sel = st.selectbox("Selecione a dívida", options=list(opcoes_div.keys()),
                                                   format_func=lambda x: opcoes_div[x],
                                                   key=f"sel_div_{mes_selecionado}")
                        with fc2:
                            val_pgto = st.number_input("Valor pago (R$)", step=10.0, min_value=0.01,
                                                       key=f"val_pgto_{mes_selecionado}")
                        with fc3:
                            st.write("")
                            pagar = st.form_submit_button("💰 Pagar")
                        if pagar and div_sel and val_pgto > 0:
                            db.adicionar_pagamento_divida(div_sel, mes_selecionado, val_pgto)
                            st.rerun()
                else:
                    st.caption("Nenhuma dívida em aberto. Cadastre abaixo ⬇️")

            # --- Lançamentos avulsos extras (sem víncilo a dívida) ---
            if lancamentos:
                st.caption("📝 Extras avulsos:")
                for l in lancamentos:
                    c1, c2, c3 = st.columns([4, 2, 1])
                    with c1:
                        novo_desc = st.text_input("Desc", l["descricao"], key=f"ld_{l['id']}", label_visibility="collapsed", disabled=_pago_mes)
                    with c2:
                        novo_val = st.number_input("Val", value=float(l["valor"]), step=10.0, key=f"lv_{l['id']}", label_visibility="collapsed", disabled=_pago_mes)
                    with c3:
                        if not _pago_mes:
                            with st.popover("🗑️"):
                                st.caption(f"Excluir **{l['descricao']}**?")
                                if st.button("✅ Confirmar", key=f"lx_ok_{l['id']}", type="primary"):
                                    db.remover_lancamento(l["id"])
                                    st.rerun()
                    if not _pago_mes and (novo_desc != l["descricao"] or novo_val != l["valor"]):
                        db.atualizar_lancamento(l["id"], novo_desc, novo_val)
                        st.rerun()

            if not _pago_mes:
                with st.form(f"add_{cat_key}_{mes_selecionado}", clear_on_submit=True):
                    st.markdown("**📝 Adicionar extra avulso:**")
                    fc1, fc2, fc3 = st.columns([4, 2, 1])
                    with fc1:
                        desc = st.text_input("Descrição do extra", key=f"nd_{cat_key}",
                                             placeholder="Ex: Farmácia, Presente...")
                    with fc2:
                        val = st.number_input("Valor (R$)", value=0.0, min_value=0.0, step=10.0,
                                              format="%.2f", key=f"nv_{cat_key}")
                    with fc3:
                        st.write("")
                        add = st.form_submit_button("➕ Add")
                    if add:
                        if not desc:
                            st.warning("Informe a descrição.")
                        elif val <= 0:
                            st.warning("Informe um valor maior que zero.")
                        else:
                            db.adicionar_lancamento(mes_selecionado, cat_key, desc, val)
                            st.rerun()

            # --- Gerenciar dívidas (cadastrar/ver) ---
            with st.expander("📋 Gerenciar Dívidas"):
                todas_dividas = db.listar_dividas(apenas_abertas=False)
                if todas_dividas:
                    for d in todas_dividas:
                        saldo = db.saldo_divida(d["id"])
                        status = "✅ Quitada" if d["quitada"] else f"Saldo: {fmt(saldo)}"
                        c1, c2, c3 = st.columns([4, 2, 1])
                        with c1:
                            st.text(f"{d['descricao']}  ({status})")
                        with c2:
                            st.text(f"Total: {fmt(d['valor_total'])}")
                        with c3:
                            if not _pago_mes:
                                with st.popover("🗑️"):
                                    st.caption(f"Excluir dívida **{d['descricao']}**?")
                                    if st.button("✅ Confirmar", key=f"dd_ok_{d['id']}", type="primary"):
                                        db.remover_divida(d["id"])
                                        st.rerun()

                if not _pago_mes:
                    with st.form(f"nova_divida_{mes_selecionado}", clear_on_submit=True):
                        st.caption("Cadastrar nova dívida:")
                        dc1, dc2, dc3 = st.columns([4, 2, 1])
                        with dc1:
                            desc_div = st.text_input("Descrição", key="nd_divida", label_visibility="collapsed",
                                                     placeholder="Ex: Troca do carburador...")
                        with dc2:
                            val_div = st.number_input("Valor total", step=10.0, min_value=0.01,
                                                      key="nv_divida", label_visibility="collapsed",
                                                      placeholder="R$")
                        with dc3:
                            add_div = st.form_submit_button("➕")
                        if add_div and desc_div and val_div > 0:
                            db.adicionar_divida(desc_div, val_div)
                            st.rerun()

            continue  # pula o bloco genérico abaixo

        # ── Categorias normais ──
        st.markdown(f'<div class="cat-header"><strong>{icon} {cat_nome}</strong> — {fmt(subtotal)}</div>',
                    unsafe_allow_html=True)

        if lancamentos:
            for l in lancamentos:
                c1, c2, c3 = st.columns([4, 2, 1])
                with c1:
                    novo_desc = st.text_input("Desc", l["descricao"], key=f"ld_{l['id']}", label_visibility="collapsed", disabled=_pago_mes)
                with c2:
                    novo_val = st.number_input("Val", value=float(l["valor"]), step=10.0, key=f"lv_{l['id']}", label_visibility="collapsed", disabled=_pago_mes)
                with c3:
                    if not _pago_mes:
                        with st.popover("🗑️"):
                            st.caption(f"Excluir **{l['descricao']}**?")
                            if st.button("✅ Confirmar", key=f"lx_ok_{l['id']}", type="primary"):
                                db.remover_lancamento(l["id"])
                                st.rerun()
                if not _pago_mes and (novo_desc != l["descricao"] or novo_val != l["valor"]):
                    db.atualizar_lancamento(l["id"], novo_desc, novo_val)
                    st.rerun()

        # Quick add inline
        if not _pago_mes:
            with st.form(f"add_{cat_key}_{mes_selecionado}", clear_on_submit=True):
                fc1, fc2, fc3 = st.columns([4, 2, 1])
                with fc1:
                    desc = st.text_input("Descrição", key=f"nd_{cat_key}", label_visibility="collapsed",
                                         placeholder=f"Nova {cat_nome.lower()}...")
                with fc2:
                    val = st.number_input("Valor (R$)", value=0.0, min_value=0.0, step=10.0,
                                          format="%.2f", key=f"nv_{cat_key}", label_visibility="collapsed")
                with fc3:
                    add = st.form_submit_button("➕")
                if add:
                    if not desc:
                        st.warning("Informe a descrição.")
                    elif val <= 0:
                        st.warning("Informe um valor maior que zero.")
                    else:
                        db.adicionar_lancamento(mes_selecionado, cat_key, desc, val)
                    st.rerun()

    # Ações do mês
    st.divider()
    col_resumo, col_acoes = st.columns([1, 1])
    with col_resumo:
        if st.button("📋 Gerar Resumo WhatsApp", use_container_width=True, type="primary"):
            st.session_state["mostrar_resumo_mes"] = mes_selecionado
    with col_acoes:
        if not _pago_mes:
            with st.expander("⚠️ Ações"):
                if st.session_state.get("_confirm_del_mes") == mes_selecionado:
                    st.error(f"⚠️ Excluir **{opcoes_mes[mes_selecionado]}** e todos os seus lançamentos? Esta ação não pode ser desfeita.")
                    cm1, cm2 = st.columns(2)
                    if cm1.button("✅ Sim, excluir", key="del_mes_ok", type="primary"):
                        db.remover_mes(mes_selecionado)
                        del st.session_state["_confirm_del_mes"]
                        st.rerun()
                    if cm2.button("❌ Cancelar", key="del_mes_no"):
                        del st.session_state["_confirm_del_mes"]
                        st.rerun()
                else:
                    if st.button("🗑️ Excluir este mês e todos os lançamentos", type="secondary"):
                        st.session_state["_confirm_del_mes"] = mes_selecionado
                        st.rerun()

    if st.session_state.get("mostrar_resumo_mes") == mes_selecionado:
        st.markdown("---")
        st.markdown("### 📋 Resumo para WhatsApp / Print")
        st.info("💡 Tire um print desta tela para enviar pelo WhatsApp com o comprovante do banco.")
        st.markdown(_resumo_mes_html(mes_selecionado), unsafe_allow_html=True)
        if st.button("✖️ Fechar Resumo", key="fechar_resumo_mes"):
            del st.session_state["mostrar_resumo_mes"]
            st.rerun()


# ═══════════════════════════════════════════════════════════════════════════
# ✈️ VIAGENS / EVENTOS
# ═══════════════════════════════════════════════════════════════════════════
elif pagina == "✈️ Viagens / Eventos":
    st.header("✈️ Viagens / Eventos Especiais")

    meses = db.listar_meses()
    viagens = db.listar_viagens()

    # Criar viagem
    with st.expander("➕ Nova Viagem / Evento"):
        with st.form("add_viagem", clear_on_submit=True):
            c1, c2, c3 = st.columns(3)
            with c1:
                nome_v = st.text_input("Nome (ex: Birigui 18-07)")
            with c2:
                data_v = st.date_input("Data", value=datetime.now(), format="DD/MM/YYYY")
            with c3:
                opcoes_mes_v = {0: "— Nenhum —"}
                for m in meses:
                    opcoes_mes_v[m["id"]] = nome_mes(m["ano"], m["mes"])
                mes_v = st.selectbox("Vincular ao mês", options=list(opcoes_mes_v.keys()),
                                     format_func=lambda x: opcoes_mes_v[x])
            if st.form_submit_button("Criar Viagem", type="primary"):
                if nome_v:
                    db.criar_viagem(nome_v, str(data_v), mes_v if mes_v else None)
                    st.rerun()

    if not viagens:
        st.info("Nenhuma viagem registrada. Crie uma acima.")
        st.stop()

    # Selecionar viagem
    opcoes_viagem = {
        v["id"]: v["nome"] + (" ✅ Pago" if ("pago" in v.keys() and v["pago"]) else "")
        for v in viagens
    }
    viagem_sel = st.selectbox("Selecione a viagem", options=list(opcoes_viagem.keys()),
                              format_func=lambda x: opcoes_viagem[x])

    v_info = db.obter_viagem(viagem_sel)
    _pago_v = bool(v_info["pago"]) if v_info and "pago" in v_info.keys() else False
    totais_v = db.totais_viagem(viagem_sel)
    total_v = db.total_viagem(viagem_sel)
    total_nina_v = db.total_nina_viagem(viagem_sel)

    # ── Banner e toggle Pago ────────────────────────────────────────────────────────
    if _pago_v:
        st.markdown("""
        <style>
        section[data-testid="stMain"] > div {
            background: rgba(40, 167, 90, 0.06) !important;
        }
        </style>""", unsafe_allow_html=True)
        st.success("✅ **VIAGEM PAGA** — Acerto concluído. Edições desativadas.")
        if st.button("🔓 Desmarcar como Pago", key="des_pago_v"):
            db.marcar_viagem_paga(viagem_sel, 0)
            st.rerun()
    else:
        if st.button("✅ Marcar como Pago", key="marcar_pago_v", type="secondary",
                     help="Bloqueia edições e sinaliza a viagem como acertada"):
            db.marcar_viagem_paga(viagem_sel, 1)
            st.rerun()

    c1, c2, c3 = st.columns(3)
    with c1:
        metric_card("Total Viagem", total_v, "blue")
    with c2:
        metric_card("Pago por Nina", total_nina_v, "orange")
    with c3:
        metric_card("Meu Custo", total_v - total_nina_v, "red")

    # Observações da viagem
    _obs_v_atual = ""
    if v_info and "observacoes" in v_info.keys():
        _obs_v_atual = v_info["observacoes"] or ""
    _nova_obs_v = st.text_input(
        "📝 Observações da viagem (opcional)",
        value=_obs_v_atual,
        placeholder="Ex: Viagem de aniversário, hotel incluso...",
        key=f"obs_viagem_{viagem_sel}",
        disabled=_pago_v,
    )
    if not _pago_v and _nova_obs_v != _obs_v_atual:
        db.atualizar_observacoes_viagem(viagem_sel, _nova_obs_v)
        st.rerun()

    st.divider()

    cat_icons_v = {
        "pedagio": "🚗", "combustivel": "⛽",
        "despesa_viagem": "🍽️", "desp_nina_viagem": "🛒",
    }

    for cat_key, cat_nome in db.CATEGORIAS_VIAGEM:
        icon = cat_icons_v.get(cat_key, "📌")
        lancamentos = db.listar_lancamentos_viagem(viagem_sel, cat_key)
        subtotal = totais_v.get(cat_key, 0)

        st.markdown(f'<div class="cat-header"><strong>{icon} {cat_nome}</strong> — {fmt(subtotal)}</div>',
                    unsafe_allow_html=True)

        if lancamentos:
            for l in lancamentos:
                c1, c2, c3, c4 = st.columns([3, 2, 1, 1])
                with c1:
                    novo_desc = st.text_input("Desc", l["descricao"], key=f"vd_{l['id']}", label_visibility="collapsed", disabled=_pago_v)
                with c2:
                    novo_val = st.number_input("Val", value=float(l["valor"]), step=10.0, key=f"vv_{l['id']}", label_visibility="collapsed", disabled=_pago_v)
                with c3:
                    nova_nina = st.checkbox("Nina", value=bool(l["pago_por_nina"]), key=f"vn_{l['id']}", disabled=_pago_v)
                with c4:
                    if not _pago_v:
                        with st.popover("🗑️"):
                            st.caption(f"Excluir **{l['descricao']}**?")
                            if st.button("✅ Confirmar", key=f"vx_ok_{l['id']}", type="primary"):
                                db.remover_lancamento_viagem(l["id"])
                                st.rerun()
                if not _pago_v and (novo_desc != l["descricao"] or novo_val != l["valor"] or nova_nina != bool(l["pago_por_nina"])):
                    db.atualizar_lancamento_viagem(l["id"], novo_desc, novo_val, nova_nina)
                    st.rerun()

        # Quick add
        if not _pago_v:
            with st.form(f"addv_{cat_key}_{viagem_sel}", clear_on_submit=True):
                fc1, fc2, fc3, fc4 = st.columns([3, 2, 1, 1])
                with fc1:
                    desc = st.text_input("Desc", key=f"nvd_{cat_key}", label_visibility="collapsed",
                                         placeholder="Nova despesa...")
                with fc2:
                    val = st.number_input("Valor (R$)", value=0.0, min_value=0.0, step=10.0,
                                          format="%.2f", key=f"nvv_{cat_key}", label_visibility="collapsed")
                with fc3:
                    nina = st.checkbox("Nina", key=f"nvn_{cat_key}")
                with fc4:
                    add = st.form_submit_button("➕")
                if add:
                    if not desc:
                        st.warning("Informe a descrição.")
                    elif val <= 0:
                        st.warning("Informe um valor maior que zero.")
                    else:
                        db.adicionar_lancamento_viagem(viagem_sel, cat_key, desc, val, nina)
                        st.rerun()

    st.divider()
    col_rv, col_av = st.columns([1, 1])
    with col_rv:
        if st.button("📋 Gerar Resumo WhatsApp", use_container_width=True, type="primary", key="btn_resumo_viagem"):
            st.session_state["mostrar_resumo_viagem"] = viagem_sel
    with col_av:
        if not _pago_v:
            with st.expander("⚠️ Ações"):
                # ── Editar nome e data ──
                _data_edit = datetime.strptime(v_info["data_viagem"], "%Y-%m-%d") if v_info["data_viagem"] else datetime.now()
                col_en, col_ed = st.columns(2)
                with col_en:
                    novo_nome_v = st.text_input("Nome da viagem", value=v_info["nome"], key=f"edit_nome_v_{viagem_sel}")
                with col_ed:
                    nova_data_v = st.date_input("Data", value=_data_edit, format="DD/MM/YYYY", key=f"edit_data_v_{viagem_sel}")
                if st.button("💾 Salvar nome/data", key=f"salvar_nome_data_v_{viagem_sel}"):
                    if novo_nome_v:
                        db.atualizar_viagem(viagem_sel, novo_nome_v, str(nova_data_v), v_info["mes_id"])
                        st.rerun()
                    else:
                        st.warning("O nome não pode ficar em branco.")
                st.divider()
                # ── Vincular ao mês ──
                meses_opcoes_edit = {0: "— Nenhum —"}
                for m_item in meses:
                    meses_opcoes_edit[m_item["id"]] = nome_mes(m_item["ano"], m_item["mes"])
                current_mes = v_info["mes_id"] if v_info["mes_id"] else 0
                novo_mes_link = st.selectbox("Vincular ao mês", options=list(meses_opcoes_edit.keys()),
                                             format_func=lambda x: meses_opcoes_edit[x],
                                             index=list(meses_opcoes_edit.keys()).index(current_mes) if current_mes in meses_opcoes_edit else 0)
                if novo_mes_link != current_mes:
                    db.atualizar_viagem(viagem_sel, v_info["nome"], v_info["data_viagem"],
                                        novo_mes_link if novo_mes_link else None)
                    st.rerun()

                if st.session_state.get("_confirm_del_viagem") == viagem_sel:
                    st.error(f"⚠️ Excluir a viagem **{v_info['nome']}** e todos os seus lançamentos? Esta ação não pode ser desfeita.")
                    cv1, cv2 = st.columns(2)
                    if cv1.button("✅ Sim, excluir", key="del_viagem_ok", type="primary"):
                        db.remover_viagem(viagem_sel)
                        del st.session_state["_confirm_del_viagem"]
                        st.rerun()
                    if cv2.button("❌ Cancelar", key="del_viagem_no"):
                        del st.session_state["_confirm_del_viagem"]
                        st.rerun()
                else:
                    if st.button("🗑️ Excluir esta viagem", type="secondary"):
                        st.session_state["_confirm_del_viagem"] = viagem_sel
                        st.rerun()

    if st.session_state.get("mostrar_resumo_viagem") == viagem_sel:
        st.markdown("---")
        st.markdown("### 📋 Resumo para WhatsApp / Print")
        st.info("💡 Tire um print desta tela para enviar pelo WhatsApp com o comprovante do banco.")
        st.markdown(_resumo_viagem_html(viagem_sel), unsafe_allow_html=True)
        if st.button("✖️ Fechar Resumo", key="fechar_resumo_viagem"):
            del st.session_state["mostrar_resumo_viagem"]
            st.rerun()


# ═══════════════════════════════════════════════════════════════════════════
# 📊 HISTÓRICO
# ═══════════════════════════════════════════════════════════════════════════
elif pagina == "📊 Histórico":
    st.header("📊 Histórico")

    tab_meses_h, tab_viagens_h = st.tabs(["📅 Acertos Mensais", "✈️ Viagens"])

    # ── Aba Acertos Mensais ──
    with tab_meses_h:
        hist = db.historico_mensal()
        if not hist:
            st.info("Nenhum dado histórico ainda.")
        else:
            dados = []
            for h in hist:
                dados.append({
                    "Mês": nome_mes(h["ano"], h["mes"]),
                    "Total": h["total"],
                })

            df = pd.DataFrame(dados)
            st.dataframe(df.style.format({"Total": "R$ {:.2f}"}), use_container_width=True, hide_index=True)

            if len(dados) > 1:
                st.subheader("📈 Evolução")
                chart_df = df.set_index("Mês")
                st.line_chart(chart_df, color="#667eea")

            # Por categoria
            st.subheader("📊 Detalhamento por Categoria")
            hist_cat = db.historico_por_categoria()
            cat_dict = dict(db.CATEGORIAS_MES)
            dados_cat = []
            for h in hist_cat:
                if h["categoria"]:
                    dados_cat.append({
                        "Mês": nome_mes(h["ano"], h["mes"]),
                        "Categoria": cat_dict.get(h["categoria"], h["categoria"]),
                        "Total": h["total"],
                    })
            if dados_cat:
                df_cat = pd.DataFrame(dados_cat)
                pivot = df_cat.pivot_table(index="Mês", columns="Categoria", values="Total", fill_value=0)
                st.dataframe(pivot.style.format("R$ {:.2f}"), use_container_width=True)
                st.bar_chart(pivot)

    # ── Aba Viagens ──
    with tab_viagens_h:
        hist_v = db.historico_viagens()
        if not hist_v:
            st.info("Nenhuma viagem registrada ainda.")
        else:
            dados_v = []
            for v in hist_v:
                data_fmt = ""
                if v["data_viagem"]:
                    try:
                        data_fmt = datetime.strptime(v["data_viagem"], "%Y-%m-%d").strftime("%d/%m/%Y")
                    except ValueError:
                        data_fmt = v["data_viagem"]
                dados_v.append({
                    "Viagem": v["nome"],
                    "Data": data_fmt,
                    "Total": v["total"],
                    "Nina pagou": v["total_nina"],
                    "Meu Custo": v["total"] - v["total_nina"],
                    "Status": "✅ Pago" if v["pago"] else "🔓 Aberto",
                })

            df_v = pd.DataFrame(dados_v)
            st.dataframe(
                df_v.style.format({
                    "Total": "R$ {:.2f}", "Nina pagou": "R$ {:.2f}", "Meu Custo": "R$ {:.2f}"
                }),
                use_container_width=True, hide_index=True,
            )

            # Métricas resumo
            total_todas_viag = sum(v["total"] for v in hist_v)
            total_nina_viag = sum(v["total_nina"] for v in hist_v)
            c1, c2, c3 = st.columns(3)
            with c1:
                metric_card("Total Gasto em Viagens", total_todas_viag, "blue")
            with c2:
                metric_card("Pago por Nina (total)", total_nina_viag, "orange")
            with c3:
                metric_card("Meu Custo Total", total_todas_viag - total_nina_viag, "red")

            if len(dados_v) > 1:
                st.subheader("📈 Evolução de Custo por Viagem")
                chart_viag = pd.DataFrame([{"Viagem": d["Viagem"], "Total": d["Total"], "Meu Custo": d["Meu Custo"]} for d in dados_v]).set_index("Viagem")
                st.bar_chart(chart_viag)


# ═══════════════════════════════════════════════════════════════════════════
# 📦 PARCELAMENTOS
# ═══════════════════════════════════════════════════════════════════════════
elif pagina == "📦 Parcelamentos":
    st.header("📦 Parcelamentos")

    tab_ativos, tab_fin = st.tabs(["✅ Ativos", "🏁 Finalizados"])

    with tab_ativos:
        parcelas = db.listar_parcelamentos(ativos=True)

        with st.expander("➕ Novo Parcelamento"):
            with st.form("add_parc", clear_on_submit=True):
                c1, c2, c3 = st.columns(3)
                with c1:
                    desc_p = st.text_input("Descrição")
                with c2:
                    val_p = st.number_input("Valor da Parcela (R$)", min_value=0.0, step=10.0)
                with c3:
                    num_p = st.number_input("Nº de Parcelas", min_value=1, value=10, step=1)
                c4, c5 = st.columns(2)
                with c4:
                    parc_atual = st.number_input("Parcela Atual", min_value=1, value=1, step=1)
                if st.form_submit_button("Adicionar", type="primary"):
                    if desc_p:
                        db.adicionar_parcelamento(desc_p, val_p, int(num_p), int(parc_atual))
                        st.rerun()

        if not parcelas:
            st.info("Nenhum parcelamento ativo.")
        else:
            total_parcelas = sum(p["valor_parcela"] for p in parcelas)
            metric_card("Total Mensal em Parcelas", total_parcelas, "orange")

            # Botão para lançar todas as parcelas ativas no mês escolhido
            meses_disp = db.listar_meses()
            if meses_disp:
                st.divider()
                st.markdown("**📅 Lançar parcelas ativas em um mês:**")
                lc1, lc2 = st.columns([3, 1])
                opcoes_mes_parc = {m["id"]: nome_mes(m["ano"], m["mes"]) for m in meses_disp}
                _mes_padrao = st.session_state.get("_mes_selecionado")
                idx_parc = list(opcoes_mes_parc.keys()).index(_mes_padrao) if _mes_padrao in opcoes_mes_parc else 0
                with lc1:
                    mes_parc = st.selectbox(
                        "Selecione o mês destino",
                        options=list(opcoes_mes_parc.keys()),
                        format_func=lambda x: opcoes_mes_parc[x],
                        index=idx_parc,
                        key="sel_mes_parc",
                    )
                with lc2:
                    st.write("")
                    if st.button("📅 Lançar no Mês", type="primary", key="btn_lancar_parc"):
                        for p in parcelas:
                            proxima = min(p["parcela_atual"] + 1, p["num_parcelas"])
                            desc_lanc = f"{p['descricao']} ({proxima}/{p['num_parcelas']})"
                            db.adicionar_lancamento(mes_parc, "parcelada", desc_lanc, p["valor_parcela"])
                            db.atualizar_parcelamento(p["id"], p["descricao"], p["valor_parcela"],
                                                      p["num_parcelas"], proxima)
                            if proxima >= p["num_parcelas"]:
                                db.finalizar_parcelamento(p["id"])
                        st.success(f"✅ {len(parcelas)} parcelamento(s) adicionado(s) em **{opcoes_mes_parc[mes_parc]}**.")
                        st.session_state["_mes_selecionado"] = mes_parc

            st.divider()
            for p in parcelas:
                progresso = p["parcela_atual"] / p["num_parcelas"]
                c1, c2, c3, c4, c5, c6, c7 = st.columns([3, 2, 2, 1, 1, 1, 1])
                with c1:
                    st.markdown(f"**{p['descricao']}**")
                with c2:
                    st.write(f"{fmt(p['valor_parcela'])}")
                with c3:
                    st.progress(progresso, text=f"{p['parcela_atual']}/{p['num_parcelas']}")
                with c4:
                    if st.button("⬅️", key=f"pb_{p['id']}", help="Voltar parcela",
                                 disabled=p["parcela_atual"] <= 1):
                        nova = max(p["parcela_atual"] - 1, 1)
                        db.atualizar_parcelamento(p["id"], p["descricao"], p["valor_parcela"],
                                                  p["num_parcelas"], nova)
                        st.rerun()
                with c5:
                    if st.button("➡️", key=f"pa_{p['id']}", help="Avançar parcela"):
                        nova = min(p["parcela_atual"] + 1, p["num_parcelas"])
                        db.atualizar_parcelamento(p["id"], p["descricao"], p["valor_parcela"],
                                                  p["num_parcelas"], nova)
                        if nova >= p["num_parcelas"]:
                            db.finalizar_parcelamento(p["id"])
                        st.rerun()
                with c6:
                    if st.button("✅", key=f"pf_{p['id']}", help="Finalizar"):
                        db.finalizar_parcelamento(p["id"])
                        st.rerun()
                with c7:
                    with st.popover("✏️", help="Editar parcelamento"):
                        st.caption(f"Editar **{p['descricao']}**")
                        _ed_desc = st.text_input("Descrição", value=p["descricao"], key=f"pe_d_{p['id']}")
                        _ed_val = st.number_input("Valor parcela", value=float(p["valor_parcela"]), step=10.0, key=f"pe_v_{p['id']}")
                        _ed_num = st.number_input("Nº parcelas", value=int(p["num_parcelas"]), min_value=1, step=1, key=f"pe_n_{p['id']}")
                        _ed_at = st.number_input("Parcela atual", value=int(p["parcela_atual"]), min_value=1, max_value=int(_ed_num), step=1, key=f"pe_a_{p['id']}")
                        if st.button("💾 Salvar", key=f"pe_ok_{p['id']}", type="primary"):
                            db.atualizar_parcelamento(p["id"], _ed_desc, _ed_val, int(_ed_num), int(_ed_at))
                            st.rerun()

    with tab_fin:
        finalizados = [p for p in db.listar_parcelamentos(ativos=False) if not p["ativo"]]
        if not finalizados:
            st.info("Nenhum parcelamento finalizado ainda.")
        else:
            st.caption(f"{len(finalizados)} parcelamento(s) concluído(s).")
            for p in finalizados:
                c1, c2, c3, c4, c5 = st.columns([3, 2, 2, 1, 1])
                with c1:
                    st.markdown(f"~~{p['descricao']}~~")
                with c2:
                    st.write(fmt(p["valor_parcela"]))
                with c3:
                    st.write(f"{p['num_parcelas']}/{p['num_parcelas']} parcelas")
                with c4:
                    if st.button("🔄", key=f"pr_{p['id']}", help="Reativar parcelamento"):
                        db.reativar_parcelamento(p["id"])
                        st.rerun()
                with c5:
                    with st.popover("🗑️", help="Excluir parcelamento"):
                        st.caption(f"Excluir **{p['descricao']}**?")
                        if st.button("✅ Confirmar", key=f"px_ok_{p['id']}", type="primary"):
                            db.remover_parcelamento(p["id"])
                            st.rerun()

            st.divider()
            if st.button("🗑️ Limpar todos os finalizados", type="secondary", key="btn_limpar_fin"):
                st.session_state["_confirm_limpar_fin"] = True
                st.rerun()
            if st.session_state.get("_confirm_limpar_fin"):
                st.error(f"⚠️ Excluir todos os {len(finalizados)} parcelamento(s) finalizado(s)? Esta ação não pode ser desfeita.")
                cf1, cf2 = st.columns(2)
                with cf1:
                    if st.button("✅ Sim, excluir todos", key="limpar_fin_ok", type="primary"):
                        for p in finalizados:
                            db.remover_parcelamento(p["id"])
                        st.session_state.pop("_confirm_limpar_fin", None)
                        st.rerun()
                with cf2:
                    if st.button("❌ Cancelar", key="limpar_fin_no"):
                        st.session_state.pop("_confirm_limpar_fin", None)
                        st.rerun()


# ═══════════════════════════════════════════════════════════════════════════
# 📥 IMPORTAR PLANILHA
# ═══════════════════════════════════════════════════════════════════════════
elif pagina == "📥 Importar / Exportar":
    st.header("📥 Importar / 📤 Exportar Dados")

    tab_exp, tab_imp, tab_email, tab_cfg = st.tabs([
        "📤 Exportar",
        "📥 Importar Planilha .xlsx",
        "📧 Restaurar do E-mail",
        "⚙️ Config E-mail",
    ])

    # ── ABA EXPORTAR ────────────────────────────────────────────────────────
    with tab_exp:
        st.caption("Gera um Excel com todos os dados e oferece download e/ou envio por e-mail.")
        cfg_exp = email_utils.get_config()
        email_ok = email_utils.is_configured(cfg_exp)
        filename_exp = f"financeiro_backup_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

        col_dl, col_mail = st.columns(2)
        with col_dl:
            if st.button("📥 Gerar e Baixar Excel", type="primary", key="btn_dl"):
                with st.spinner("Gerando arquivo..."):
                    excel_bytes = _gerar_excel_bytes()
                st.download_button(
                    label="⬇️ Clique aqui para baixar",
                    data=excel_bytes,
                    file_name=filename_exp,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_btn",
                )
                st.success("Pronto!")

        with col_mail:
            if not email_ok:
                st.warning("⚠️ Configure o e-mail na aba ⚙️ Config E-mail para habilitar.")
            else:
                if st.button("📧 Enviar Backup por E-mail", type="secondary", key="btn_send_mail",
                             help=f"Envia para: {cfg_exp.get('email_destino', '')}"):
                    with st.spinner("Enviando e-mail..."):
                        try:
                            excel_bytes = _gerar_excel_bytes()
                            email_utils.send_backup(excel_bytes, filename_exp, cfg_exp)
                            removed = email_utils.delete_old_backups(cfg_exp)
                            msg = f"✅ Backup enviado para **{cfg_exp['email_destino']}**"
                            if removed > 0:
                                msg += f" ({removed} backup(s) antigo(s) removido(s))"
                            st.success(msg)
                        except Exception as e:
                            st.error(f"❌ Erro ao enviar: {e}")

        st.divider()
        st.caption("ℹ️ Dica: use **Enviar por E-mail** antes de períodos de inatividade do deploy gratuito. "
                   "O backup fica na sua caixa de entrada e pode ser restaurado a qualquer momento.")

    # ── ABA IMPORTAR PLANILHA ORIGINAL ──────────────────────────────────────
    with tab_imp:
        st.caption(
            "Selecione o arquivo Excel para importar (ex: 'PLANEJAMENTO APOSENTADORIA.xlsx'). "
            "Útil para restaurar dados a partir da planilha original quando o backup por e-mail não for suficiente."
        )

        if not db.banco_vazio():
            st.warning(
                "⚠️ **O banco já contém dados.** A importação pode gerar duplicatas. "
                "Recomendamos usar apenas com o banco vazio (limpe via 📧 Restaurar do E-mail)."
            )
            _confirmar_imp = st.checkbox("✅ Estou ciente e desejo importar mesmo assim", key="ack_imp_dup")
        else:
            _confirmar_imp = True

        arquivo_upload = st.file_uploader(
            "Selecione o arquivo Excel (.xlsx)",
            type=["xlsx"],
            key="uploader_imp_orig",
        )

        if st.button("🚀 Importar Planilha", type="primary", key="btn_imp_orig",
                     disabled=arquivo_upload is None or not _confirmar_imp):
            import openpyxl as _openpyxl
            import io as _io

            wb = _openpyxl.load_workbook(_io.BytesIO(arquivo_upload.read()), data_only=True)
            log = []

            ws_pl = wb["Planilha2"]
            renda_map = {
                "ALLIED": (ws_pl["I2"].value or 0),
                "FIIs": (ws_pl["I3"].value or 0),
                "RENDA FIXA": (ws_pl["I4"].value or 0),
            }
            for nome, valor in renda_map.items():
                db.adicionar_fonte_renda(nome, float(valor))
                log.append(f"✅ Renda: {nome} = R$ {valor}")

            desp_rows = [
                (ws_pl[f"A{r}"].value, ws_pl[f"B{r}"].value,
                 " | ".join(filter(None, [
                     str(ws_pl[f"C{r}"].value or ""),
                     str(ws_pl[f"D{r}"].value or ""),
                     str(ws_pl[f"E{r}"].value or ""),
                 ])))
                for r in range(2, 12)
                if ws_pl[f"A{r}"].value is not None and ws_pl[f"B{r}"].value is not None
            ]
            for valor, nome, det in desp_rows:
                det_clean = det.strip(" |") if det.strip(" |") else None
                db.adicionar_despesa_planejamento(nome, float(valor), det_clean)
                log.append(f"✅ Despesa: {nome} = R$ {valor}")

            meses_map = {
                "CONTAS JULHO 2025": (2025, 7), "CONTAS AGOSTO 2025": (2025, 8),
                "CONTAS SETEMBRO 2025": (2025, 9), "CONTAS OUTUBRO 2025": (2025, 10),
                "CONTAS NOVEMBRO 2025": (2025, 11), "CONTAS DEZEMBRO 2025": (2025, 12),
                "CONTAS JANEIRO 2026": (2026, 1), "CONTAS FEVEREIRO 2026": (2026, 2),
                "CONTAS MARÇO 2026": (2026, 3),
            }
            for sheet_name, (ano, mes_num) in meses_map.items():
                if sheet_name not in wb.sheetnames:
                    continue
                ws = wb[sheet_name]
                m = db.criar_mes(ano, mes_num)
                mes_id = m["id"]
                current_cat = None
                cat_map = {
                    "contas fixas": "conta_fixa", "parceladas": "parcelada",
                    "desp. nina (devolver)": "desp_nina", "desp. nina birigua": "desp_nina",
                    "devedor (nina)": "devedor_nina",
                    "extras devedor (nino)": "extras_nino", "extras devedor(nino)": "extras_nino",
                }
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=2, max_col=3, values_only=False):
                    cell_b = row[0].value
                    cell_c = row[1].value if len(row) > 1 else None
                    if cell_b and isinstance(cell_b, str):
                        b_lower = cell_b.strip().lower()
                        if b_lower in cat_map:
                            current_cat = cat_map[b_lower]
                            continue
                        if b_lower.startswith("_") or b_lower in ("tl", "arred", "extrato :"):
                            continue
                    if current_cat and cell_b and cell_c is not None and isinstance(cell_c, (int, float)):
                        db.adicionar_lancamento(mes_id, current_cat, str(cell_b), float(cell_c))
                log.append(f"✅ Mês: {nome_mes(ano, mes_num)}")

            viagens_data = {
                "BIRIGUI 18-07": ("2025-07-18", (2025, 7)),
                "JAMPA 04-08": ("2025-08-04", (2025, 8)),
                "BIRIGUI 13-02": ("2026-02-13", (2026, 2)),
            }
            for sheet_name, (data_v, (ano_v, mes_v)) in viagens_data.items():
                if sheet_name not in wb.sheetnames:
                    continue
                ws = wb[sheet_name]
                mes_ref = db.criar_mes(ano_v, mes_v)
                db.criar_viagem(sheet_name, data_v, mes_ref["id"])
                viagens_list = db.listar_viagens()
                viagem_id = viagens_list[0]["id"]
                current_cat = None
                cat_map_v = {
                    "contas fixas": "pedagio", "combustível": "combustivel",
                    "desp birigui": "despesa_viagem", "desp jampa": "despesa_viagem",
                    "desp. nina (devolver)": "desp_nina_viagem",
                }
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=2, max_col=4, values_only=False):
                    cell_b = row[0].value
                    cell_c = row[1].value if len(row) > 1 else None
                    cell_d = row[2].value if len(row) > 2 else None
                    if cell_b and isinstance(cell_b, str):
                        b_lower = cell_b.strip().lower()
                        if b_lower in cat_map_v:
                            current_cat = cat_map_v[b_lower]
                            continue
                        if b_lower.startswith("_") or b_lower in ("tl", "arred", "extrato :"):
                            continue
                    if current_cat == "combustivel" and cell_c is not None and isinstance(cell_c, (int, float)) and not cell_b:
                        pago_nina = (cell_d == "N") if cell_d else False
                        db.adicionar_lancamento_viagem(viagem_id, "combustivel", "Combustível", float(cell_c), pago_nina)
                        continue
                    if current_cat and cell_b and cell_c is not None and isinstance(cell_c, (int, float)):
                        pago_nina = (cell_d == "N") if cell_d else False
                        db.adicionar_lancamento_viagem(viagem_id, current_cat, str(cell_b), float(cell_c), pago_nina)
                log.append(f"✅ Viagem: {sheet_name}")

            db.adicionar_parcelamento("Seguro", 200, 10, 6)
            log.append("✅ Parcelamento: Seguro 6/10")
            st.success("Importação concluída!")
            for item in log:
                st.write(item)
            st.info(
                "📧 **Backup automático garantido:** ao clicar em **Encerrar Sessão**, "
                "o app detectará que os dados mudaram (comparação por hash) e enviará "
                "automaticamente um novo backup por e-mail, sobrescrevendo o anterior. "
                "Assim você poderá restaurar esta versão importada no futuro."
            )

    # ── ABA RESTAURAR DO E-MAIL ──────────────────────────────────────────────
    with tab_email:
        cfg_restore = email_utils.get_config()
        if not email_utils.is_configured(cfg_restore):
            st.warning("⚠️ Configure o e-mail na aba ⚙️ Config E-mail para usar esta funcionalidade.")
        else:
            st.info(
                f"O app irá buscar na caixa de entrada **{cfg_restore.get('smtp_user', '')}** "
                "o e-mail de backup mais recente, **limpar todo o banco atual** e "
                "restaurar os dados do backup."
            )
            st.error(
                "🚨 **Atenção:** todos os dados atuais serão apagados antes de importar o backup. "
                "Esta ação não pode ser desfeita."
            )
            confirmar = st.checkbox("✅ Confirmo que desejo apagar os dados atuais e restaurar o backup")

            if st.button("📧 Buscar e Restaurar Backup mais Recente", type="primary",
                         key="btn_restore", disabled=not confirmar):
                with st.spinner("Conectando ao e-mail e buscando backup..."):
                    try:
                        excel_bytes, fname = email_utils.get_latest_backup(cfg_restore)
                    except Exception as e:
                        st.error(f"❌ Erro ao acessar e-mail: {e}")
                        st.stop()

                if excel_bytes is None:
                    st.error("❌ Nenhum backup encontrado na caixa de entrada. Verifique se já enviou um backup por e-mail.")
                else:
                    st.success(f"📥 Backup encontrado: **{fname}**")
                    with st.spinner("Limpando banco e restaurando dados..."):
                        import io as _io
                        import openpyxl as _opx
                        db.limpar_banco()
                        wb_bk = _opx.load_workbook(_io.BytesIO(excel_bytes), data_only=True)
                        try:
                            log_bk = _importar_backup_workbook(wb_bk)
                            st.success("✅ Restauração concluída!")
                            for item in log_bk:
                                st.write(item)
                        except Exception as e:
                            db.limpar_banco()  # garante que não fica dado pela metade
                            st.error(f"❌ Erro durante importação. Banco foi limpo: {e}")

    # ── ABA CONFIG E-MAIL ────────────────────────────────────────────────────
    with tab_cfg:
        st.caption(
            "Configure as credenciais de e-mail para envio (SMTP) e recebimento (IMAP). "
            "Para Gmail, use uma [Senha de App](https://myaccount.google.com/apppasswords) — requer 2FA ativado."
        )

        cfg_atual = email_utils.load_config()
        secrets_ativo = bool(email_utils._secrets_config().get("smtp_user"))

        if secrets_ativo:
            st.success("✅ Configurações carregadas via **st.secrets** (Streamlit Cloud). Não é possível editar aqui.")
            with st.expander("📖 Ver formato do secrets.toml"):
                st.code("""
[email]
smtp_host     = "smtp.gmail.com"
smtp_port     = 587
smtp_user     = "seu@gmail.com"
smtp_password = "xxxx xxxx xxxx xxxx"  # Senha de App
imap_host     = "imap.gmail.com"
imap_port     = 993
email_destino = "seu@gmail.com"

[auth]
username       = "Krauss"
password_hash  = "hash_sha256_da_sua_senha"
recovery_email = "krauss.christian@gmail.com"
""", language="toml")
        else:
            with st.form("form_email_cfg"):
                st.markdown("**Servidor de Envio (SMTP)**")
                c1, c2 = st.columns([3, 1])
                with c1:
                    smtp_host = st.text_input("Host SMTP", value=cfg_atual.get("smtp_host", "smtp.gmail.com"))
                with c2:
                    smtp_port = st.number_input("Porta", value=int(cfg_atual.get("smtp_port", 587)), step=1)

                st.markdown("**Credenciais**")
                c3, c4 = st.columns(2)
                with c3:
                    smtp_user = st.text_input("E-mail (usuário)", value=cfg_atual.get("smtp_user", ""))
                with c4:
                    smtp_pass = st.text_input("Senha de App", value=cfg_atual.get("smtp_password", ""),
                                              type="password")

                st.markdown("**Servidor de Recebimento (IMAP)**")
                c5, c6 = st.columns([3, 1])
                with c5:
                    imap_host = st.text_input("Host IMAP", value=cfg_atual.get("imap_host", "imap.gmail.com"))
                with c6:
                    imap_port = st.number_input("Porta IMAP", value=int(cfg_atual.get("imap_port", 993)), step=1)

                email_dest = st.text_input("E-mail de destino do backup",
                                           value=cfg_atual.get("email_destino", smtp_user),
                                           help="Pode ser o mesmo e-mail ou outro de sua preferência.")

                if st.form_submit_button("💾 Salvar Configuração", type="primary"):
                    nova_cfg = {
                        "smtp_host": smtp_host, "smtp_port": int(smtp_port),
                        "smtp_user": smtp_user, "smtp_password": smtp_pass,
                        "imap_host": imap_host, "imap_port": int(imap_port),
                        "email_destino": email_dest or smtp_user,
                    }
                    email_utils.save_config(nova_cfg)
                    st.success("✅ Configuração salva em email_config.json")
                    st.rerun()

            st.divider()
            with st.expander("📖 Usar no Streamlit Cloud (st.secrets)"):
                st.markdown(
                    "No deploy, vá em **App settings → Secrets** e adicione:"
                )
                st.code("""
[email]
smtp_host     = "smtp.gmail.com"
smtp_port     = 587
smtp_user     = "seu@gmail.com"
smtp_password = "xxxx xxxx xxxx xxxx"
imap_host     = "imap.gmail.com"
imap_port     = 993
email_destino = "seu@gmail.com"
""", language="toml")
